import os
import xlsxwriter
import openpyxl
from openpyxl import load_workbook
from plotly.subplots import make_subplots
import plotly.graph_objects as go
import plotly.express as px
import plotly.io as pio
import sklearn.metrics
import numpy as np
import pandas as pd
import warnings
warnings.simplefilter(action='ignore', category=FutureWarning)
from src.metrics_and_losses import per_prop_r2, per_prop_mse, per_prop_stable_mse
from src.dataset import DatasetProcessor
import datetime

def performance_analysis(task_type:str, artefact_dir:os.PathLike, prop_list:list, feature_rows:pd.DataFrame, 
                         block_pred:np.ndarray, block_target:np.ndarray, suffix:str, normalized:bool, fold_id:int):
    features_columns=pd.DataFrame(feature_rows)
    xlsx_name=f'outlier_analysis_{suffix}_{fold_id+1}.xlsx'
    xlsx_path=os.path.join(artefact_dir,xlsx_name) if artefact_dir else xlsx_name
    if len(features_columns)!=len(block_pred):return
    if task_type=="regression":
        with pd.ExcelWriter(xlsx_path,engine='xlsxwriter') as writer:
            for i,property_name in enumerate(prop_list):
                temp_df=features_columns.copy()
                temp_df['target']=block_target[:,i]
                temp_df['pred']=block_pred[:,i]
                temp_df['SE']=(temp_df['target']-temp_df['pred'])**2
                temp_df.reset_index(drop=True,inplace=True)
                temp_df['XXX']=np.nan
                temp_df['Categories']=np.nan
                temp_df['# Values']=np.nan
                temp_df['Category MSE']=np.nan
                temp_df['Category total SE']=np.nan
                grouped=temp_df.groupby(temp_df['Family'].str.lower())
                sums=grouped['SE'].sum()
                means=grouped['SE'].mean()
                temp_df.loc[:len(sums.index)-1,'Categories']=sums.index
                temp_df.loc[:len(sums.index)-1,'Category MSE']=means.values
                temp_df.loc[:len(sums.index)-1,'Category total SE']=sums.values
                temp_df.loc[:len(sums.index)-1,'# Values']=grouped.size().values
                sorted_outlier_analytics=temp_df.iloc[:,-4:].sort_values(by=temp_df.columns[-2],ascending=False).reset_index(drop=True)
                temp_df=pd.concat([temp_df.drop(temp_df.columns[-4:],axis=1),sorted_outlier_analytics],axis=1)
                temp_df.to_excel(writer,sheet_name=property_name,index=False)
        outlier_scatters(artefact_dir,suffix,xlsx_path,normalized,fold_id)
        if not normalized:print(f'{suffix} outlier analysis complete!')
    else:
        roc_auc_results = []

        for i, property_name in enumerate(prop_list):
            df = pd.DataFrame({
                'target': block_target[:, i],
                'pred': block_pred[:, i]
            })
            df = df.dropna(subset=['target'])

            if df['target'].nunique() != 2 or len(df) < 2:
                auc_value = np.nan  
            else:
                try:
                    auc_value = sklearn.metrics.roc_auc_score(df['target'], df['pred'])
                except ValueError:
                    auc_value = np.nan  

            roc_auc_results.append([property_name, auc_value])

        df_auc = pd.DataFrame(roc_auc_results, columns=["Property", "ROC-AUC"])

        if df_auc["ROC-AUC"].notna().sum() > 0:  
            with pd.ExcelWriter(xlsx_path, engine='xlsxwriter') as writer:
                df_auc.to_excel(writer, sheet_name="Classification_Metrics", index=False)
            print(f"Classification ROC-AUC analysis file written: {xlsx_path}")
            
def outlier_scatters(artefact_dir, suffix, xlsx_path, normalized, fold_id):
    sheets=pd.read_excel(xlsx_path,sheet_name=None)
    property_full_names=DatasetProcessor.property_full_names
    all_r2=[]
    fig=make_subplots(rows=8,cols=8,subplot_titles=[property_full_names.get(k,k) for k in sheets.keys()])
    color_map=px.colors.qualitative.Plotly
    for i,(sheet_name,sheet_data) in enumerate(sheets.items(),start=1):
        row=(i-1)//8+1
        col=(i-1)%8+1
        sheet_data=sheet_data.dropna(subset=['target']).copy()
        sheet_data.loc[:,'color']=sheet_data.loc[:,'Family'].apply(lambda x:color_map[hash(str(x))%len(color_map)])
        text=[f"Family: {fam}<br>SMILES: {sm}" for fam,sm in zip(sheet_data['Family'].tolist(),sheet_data['SMILES'].tolist())]
        scatter=go.Scatter(x=sheet_data['target'],y=sheet_data['pred'],mode='markers',marker=dict(size=5,color=sheet_data['color'],showscale=False),text=text,hovertemplate='%{text}<br>Target: %{x}<br>Prediction: %{y}')
        fig.add_trace(scatter,row=row,col=col)
        if len(sheet_data['target'])>0:
            fig.add_trace(go.Scatter(x=[min(sheet_data['target']),max(sheet_data['target'])],y=[min(sheet_data['target']),max(sheet_data['target'])],mode='lines',line=dict(color='rgba(0, 0, 0, 0.5)',width=1),showlegend=False),row=row,col=col)
        r2=per_prop_r2(np.array(sheet_data['pred'])[:,np.newaxis],np.array(sheet_data['target'])[:,np.newaxis],False)[0]
        r2_stable=per_prop_r2(np.array(sheet_data['pred'])[:,np.newaxis],np.array(sheet_data['target'])[:,np.newaxis],True)[0]
        mse=per_prop_mse(np.array(sheet_data['pred'])[:,np.newaxis],np.array(sheet_data['target'])[:,np.newaxis])[0]
        stable_mse=per_prop_stable_mse(np.array(sheet_data['pred'])[:,np.newaxis],np.array(sheet_data['target'])[:,np.newaxis])[0]
        fig.update_xaxes(title_text='Target',row=row,col=col)
        fig.update_yaxes(title_text='Prediction',row=row,col=col)
        if normalized:
            title=property_full_names.get(sheet_name,sheet_name)+f" (n={len(sheet_data['target'])})<br>R² = {r2:.3f}  R² stable = {r2_stable:.3f}<br>MSE = {mse:.3f}  MSE stable = {stable_mse:.3f}"
        else:
            title=property_full_names.get(sheet_name,sheet_name)+f" (n={len(sheet_data['target'])})<br>R² = {r2:.3f}  R² stable = {r2_stable:.3f}<br>RMSE = {np.sqrt(mse):.3e}  RMSE stable = {np.sqrt(stable_mse):.3e}"
        fig.layout.annotations[i-1].update(text=title,font=dict(size=11))
        all_r2.append(r2)
    fig.update_layout(showlegend=False)
    avg_r2=np.mean(all_r2) if all_r2 else 0.0
    fig.update_layout(title=f"<b>Average R² across all tasks = {avg_r2:.4f}</b>",title_x=0.5,title_y=0.98)
    img_name=f"interactive_normalized_scatters_{suffix}_{fold_id+1}.html" if normalized else f"interactive_scatters_{suffix}_{fold_id+1}.html"
    img_path=os.path.join(artefact_dir,img_name) if artefact_dir else img_name
    pio.write_html(fig,img_path)

def k_fold_analysis(datetime_str:datetime,limit_folds:int):
    run_directory=training_directory_finder(datetime_str)
    targets=load_workbook(os.path.join(run_directory,'fold_1','outlier_analysis_val_1.xlsx'),read_only=True).sheetnames
    rows=[]
    with pd.ExcelWriter(os.path.join(run_directory,'k_fold_performance.xlsx'),engine='xlsxwriter') as writer:
        for mode in['val','train_for_eval', 'test']:
            for target in targets:
                per_fold_rmse_list=[]
                per_fold_r2_list=[]
                for fold in range(limit_folds):
                    performance_df=pd.read_excel(os.path.join(run_directory,f"fold_{fold+1}",f'outlier_analysis_{mode}_{fold+1}.xlsx'),sheet_name=target)
                    per_fold_rmse_list.append(rmse_extractor(performance_df))
                    per_fold_r2_list.append(r2_extractor(performance_df))
                for metric in['R2','RMSE']:
                    if metric=='R2':
                        rows.append([target,metric]+[per_fold_r2_list[fold] for fold in range(limit_folds)])
                    else:
                        rows.append([target,metric]+[per_fold_rmse_list[fold] for fold in range(limit_folds)])
            columns=['Target','Metric']+[f'Fold {i+1}' for i in range(limit_folds)]
            output_df=pd.DataFrame(rows,columns=columns)
            fold_data=output_df.iloc[:,-limit_folds:]
            output_df['Mean']=fold_data.mean(axis=1)
            output_df['STD']=fold_data.std(axis=1)
            if mode=='val':
                avg_r2=np.mean(output_df['Mean'][output_df['Metric']=='R2'])
                output_df['Overall R²']=[avg_r2]+[np.nan]*(len(output_df)-1)
                output_df.to_excel(writer,sheet_name='Val',index=False)
            else:
                avg_r2=np.mean(output_df['Mean'][output_df['Metric']=='R2'])
                output_df['Overall R²']=[avg_r2]+[np.nan]*(len(output_df)-1)
                output_df.to_excel(writer,sheet_name='Train',index=False)
            del rows[:]

def rmse_extractor(performance_df:pd.DataFrame):
    mask=np.isnan(performance_df['target'])
    performance_df=performance_df[~mask]
    rmse=np.sqrt(np.mean(performance_df['SE']))
    return rmse

def r2_extractor(performance_df:pd.DataFrame):
    mask=np.isnan(performance_df['target'])
    performance_df=performance_df[~mask]
    if len(performance_df['target'])<=2:return np.nan
    else:return sklearn.metrics.r2_score(performance_df['target'],performance_df['pred'])

def training_directory_finder(datetime_str:datetime):
    runs_list=[name for name in os.listdir('runs') if os.path.isdir(os.path.join('runs',name))]
    idx=next((i for i,entry in enumerate(runs_list) if entry[:19]==datetime_str),-1)
    run_directory=os.path.join('runs',runs_list[idx])
    return run_directory
